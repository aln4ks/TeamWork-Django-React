import datetime
from itertools import chain
from urllib.parse import quote

from openpyxl import Workbook
from openpyxl.utils import get_column_letter
from django.http import HttpResponse
from rest_framework.views import APIView

from teamwork.views.auxiliary import auth_required
from teamwork.models import Project, Task, STATUS_CHOICES, CATEGORY_CHOICES


class ExportToXLSXView(APIView):
    @staticmethod
    def _get_project(project_id):
        return Project.objects.get(id=project_id)

    @staticmethod
    def _get_project_data(project):
        return {
            "Название проекта": project.name,
            "Описание проекта": project.description,
            "Кол-во участников проекта": project.employee.count(),
            "Дата создания проекта": project.created_at.strftime("%Y-%m-%d %H:%M"),
            "Создатель проекта": project.get_creator_name(),
            "Дата экспорта": datetime.datetime.now().strftime("%Y-%m-%d %H:%M")
        }

    @staticmethod
    def _get_task_data(task):
        return {
            "Название": task.name,
            "Описание": task.description,
            "Исполнитель": f"{task.executor.first_name} {task.executor.last_name}",
            "Создатель": f"{task.creator.first_name} {task.creator.last_name}",
            "Приоритет": task.priority,
            "Категория": task.category,
            "Дата создания": task.created_at.strftime("%Y-%m-%d %H:%M"),
            "Статус": task.status,
            "Дэдлайн": task.deadline.strftime("%Y-%m-%d %H:%M"),
            "Дата выполнения": task.completed_date.strftime("%Y-%m-%d %H:%M") if task.completed_date else None,
            "Удалена": "Да" if task.is_deleted and task.status != STATUS_CHOICES[-1][0] else "Нет",
            "Дата удаления": task.deleted_at.strftime("%Y-%m-%d %H:%M") if task.is_deleted else None
        }

    @staticmethod
    def auto_column_size(ws):
        for col in range(1, ws.max_column + 1):
            max_length = 0
            column = get_column_letter(col)

            if ws[f"{column}1"].value:
                max_length = max(max_length, len(str(ws[f"{column}1"].value)))

            for cell in ws[f"{column}"][1:]:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass

            ws.column_dimensions[column].width = max_length + 5

    @staticmethod
    def get_tasks_result(tasks):
        bags = [task for task in tasks if task.category == CATEGORY_CHOICES[0][0]]
        completed_tasks = [task for task in tasks if task.status == STATUS_CHOICES[-1][0]]
        completed_on_time_tasks = [task for task in completed_tasks if task.deadline >= task.completed_date]
        deleted_tasks = [task for task in tasks if task.is_deleted and task not in completed_tasks]
        results = {
            "Итого задач": len(tasks),
            "Багов": len(bags),
            "Удалено": len(deleted_tasks),
            "Выполнено": len(completed_tasks),
            "Выполнено в срок": len(completed_on_time_tasks)
        }
        return results

    @auth_required
    def get(self, request, project_id):
        project = self._get_project(project_id)
        project_data = self._get_project_data(project)

        active_tasks = Task.objects.filter(project=project).select_related("creator", "executor")
        deleted_tasks = Task.deleted_objects.filter(project=project).select_related("creator", "executor")
        tasks_data = []
        for task in chain(active_tasks, deleted_tasks):
            tasks_data.append(self._get_task_data(task))

        wb = Workbook()
        ws = wb.active
        ws.title = "Отчет по проекту"

        header_row = 1
        data_row = 2

        for col, key in enumerate(project_data.keys(), start=1):
            ws.cell(header_row, col, key)

        for col, value in enumerate(project_data.values(), start=1):
            ws.cell(data_row, col, value)

        header_row += 3
        data_row += 3
        ws.cell(header_row, 1, "Задачи проекта:")
        header_row += 1
        data_row += 1

        for col, key in enumerate(tasks_data[0].keys(), start=1):
            ws.cell(header_row, col, key)

        for task in tasks_data:
            for col, value in enumerate(task.values(), start=1):
                ws.cell(data_row, col, value)
            data_row += 1

        results = self.get_tasks_result(active_tasks.union(deleted_tasks))

        for row, (key, value) in enumerate(results.items(), start=data_row + 1):
            ws.cell(row, 1, key)
            ws.cell(row, 2, value)

        self.auto_column_size(ws)

        response = HttpResponse(content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")
        report_name = f"{project.name.replace(' ', '_')}_report.xlsx"
        response["Content-Disposition"] = f'attachment; filename="{quote(report_name)}"'
        wb.save(response)
        return response
